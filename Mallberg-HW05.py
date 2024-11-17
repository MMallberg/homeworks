#Marit Mallberg
#cosc 1010
#Lab Section 10
#Help given/recieved:
#Notes: I decided to make a geometric mandela/snowflake pattern, it is of my own creation. 

import openpyxl
from openpyxl.styles import Color, PatternFill

import string

canvas = openpyxl.workbook()
sheet = canvas.active

#set cells as squares ...
for i in range(1,24):
    sheet.row_demensions[i].height = 23.25
for chr in string.ascii_uppercase[:21]:
    sheet.column_dimensions[chr].width = 3.71

# set colors and fills... 
greyish = Color(rgb='B5D6B2')
greyfill = PatternFill(patternType = 'solid', fgColor=greyish)

teal = Color(rgb='086375')
tealfill = PatternFill(patternType='solid', fgColor=teal)

purple = Color(rgb='3D5B26')
purplefill = PatternFill(patternType='solid', fgColor=purple)

sky = Color(rgb='8093F1')
skyfill = PatternFill(patternType='solid', fgColor= sky)

green = Color(rgb='397367')
greenfill = PatternFill(patternType='solid',fgColor=green)

#now to set a list to each color
tealcells = ["H11","H12","H13","I12","J12","K12","L12","M12","N11","N12","N13"]
purplecells = ["K8","K9","K10","J9","L9","K14","K15","K16","J15","L15"]
greycells = ["C12","D12","E11","E13","F11","F13","G10","G14","H9","H15","I7","I8","I16","I17","J6","J18","K5","K18","L6","L18","M7","M8","M16","M17","N9","N15","O8","O14","P11","P13","Q11","Q13","R12","S12"]
skycells = ["E6","E7","E17","E18","F6","F7","F17","F18","G8","G16","O8","O16","P6","P7","P17","P18","Q6","Q7","Q17","Q18"]
greencells = ["A12","B11","B13","I3","I21","J4","J20","K1","K4","K20","K23","L4","L20","M3","M21",]

parameter_check = len(tealcells)+len(purplecells)+len(greycells)+len(skycells)+len(greencells)
print(parameter_check)

#Now to color ... 

for i in range(1,24):
    for chr in string.ascii_uppercase[:21]:
        coordin = chr+string(i)
        if coordin in tealcells:
            sheet[coordin].fill = tealfill
        if coordin in purplecells:
            sheet[coordin].fill = purplecells
        if coordin in greycells:
            sheet[coordin].fill = greyfill
        if coordin in skycells:
            sheet[coordin].fill = skyfill
        if coordin in greencells:
            sheet[coordin].fill = greenfill

canvas.save('HW05-cosc.xlsx')